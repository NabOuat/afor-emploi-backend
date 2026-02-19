# -*- coding: utf-8 -*-
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from app.database import get_db
from app.models import FicPersonne, Contrat, FicPersonneLocalisation, Acteur, Projet, Engagement, ProjetEngagement
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
import uuid
from datetime import datetime

router = APIRouter(prefix="/api/import-export", tags=["import-export"])


@router.get("/download-template")
async def download_template():
    """Télécharger le template Excel pour l'import d'employés"""
    
    try:
        # Créer un nouveau classeur
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Employés"
        
        # Définir les colonnes
        headers = [
            "Nom",
            "Prénom",
            "Date de Naissance (YYYY-MM-DD)",
            "Genre (M/F)",
            "Contact",
            "Projet",
            "Engagement",
            "Poste Nom",
            "Catégorie Poste",
            "Type Contrat (CDI/CDD/Stage)",
            "Type Personne",
            "Poste",
            "Diplôme",
            "École",
            "Date Début Contrat (YYYY-MM-DD)",
            "Date Fin Contrat (YYYY-MM-DD)",
            "Région",
            "Département",
            "Sous-Préfecture"
        ]
        
        # Ajouter les headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Ajouter une ligne d'exemple
        example_data = [
            "Dupont",
            "Jean",
            "1990-05-15",
            "M",
            "+225 01 23 45 67",
            "Projet Cartographie",
            "ETAT",
            "Opérateur Topographe",
            "Technique",
            "CDI",
            "Permanent",
            "MAITRISE EN GEOGRAPHIE",
            "Bac+2",
            "Institut Géomatique",
            "2024-01-15",
            "2025-01-15",
            "Abidjan",
            "Abidjan",
            "Abidjan"
        ]
        
        for col_num, value in enumerate(example_data, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Ajuster les largeurs de colonnes
        column_widths = [15, 15, 20, 10, 20, 20, 15, 20, 20, 18, 15, 20, 15, 20, 20, 20, 15, 15, 20]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width
        
        # Ajouter une feuille d'instructions
        instructions_ws = wb.create_sheet("Instructions")
        instructions = [
            ["INSTRUCTIONS D'IMPORT"],
            [""],
            ["1. Remplissez la feuille 'Employés' avec les données des employés"],
            ["2. Les colonnes obligatoires sont : Nom, Prénom, Projet, Poste Nom, Date Début Contrat"],
            ["3. Les dates doivent être au format YYYY-MM-DD (ex: 2024-01-15)"],
            ["4. Le Genre doit être M ou F"],
            ["5. Type Contrat doit être : CDI, CDD, Stage, etc."],
            ["6. Engagement est OPTIONNEL - laissez vide si non applicable"],
            ["7. Les colonnes de localisation (Région, Département, Sous-Préfecture) sont optionnelles"],
            ["8. Sauvegardez le fichier en format Excel (.xlsx)"],
            ["9. Importez le fichier via le bouton 'Importer Excel'"],
            [""],
            ["COLONNES DISPONIBLES:"],
            ["- Nom: Nom de l'employé (obligatoire)"],
            ["- Prénom: Prénom de l'employé (obligatoire)"],
            ["- Date de Naissance: Format YYYY-MM-DD (optionnel)"],
            ["- Genre: M ou F (optionnel)"],
            ["- Contact: Numéro de téléphone ou email (optionnel)"],
            ["- Projet: Nom du projet (obligatoire)"],
            ["- Engagement: Engagement du projet (OPTIONNEL) - ex: ETAT, IDA"],
            ["- Poste Nom: Nom du poste (obligatoire)"],
            ["- Catégorie Poste: Catégorie du poste (optionnel)"],
            ["- Type Contrat: CDI, CDD, Stage, etc. (optionnel)"],
            ["- Type Personne: Type de personne (optionnel)"],
            ["- Poste: Poste détaillé (optionnel)"],
            ["- Diplôme: Niveau d'éducation (optionnel)"],
            ["- École: École/Formation (optionnel)"],
            ["- Date Début Contrat: Format YYYY-MM-DD (obligatoire)"],
            ["- Date Fin Contrat: Format YYYY-MM-DD (optionnel)"],
            ["- Région: Région de localisation (optionnel)"],
            ["- Département: Département de localisation (optionnel)"],
            ["- Sous-Préfecture: Sous-préfecture de localisation (optionnel)"],
        ]
        
        for row_num, row_data in enumerate(instructions, 1):
            for col_num, value in enumerate(row_data, 1):
                cell = instructions_ws.cell(row=row_num, column=col_num)
                cell.value = value
                if row_num == 1:
                    cell.font = Font(bold=True, size=14)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(bold=True, size=14, color="FFFFFF")
        
        instructions_ws.column_dimensions['A'].width = 80
        
        # Sauvegarder dans BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=template_employes_{datetime.now().strftime('%Y%m%d')}.xlsx"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/import-employees")
async def import_employees(
    file: UploadFile = File(...),
    acteur_id: str = None,
    projet_id: str = None,
    db: Session = Depends(get_db)
):
    """Importer des employés depuis un fichier Excel"""
    
    try:
        # Lire le fichier Excel
        contents = await file.read()
        wb = openpyxl.load_workbook(BytesIO(contents))
        ws = wb.active
        
        results = {
            "total": 0,
            "success": 0,
            "errors": [],
            "employees": []
        }
        
        # Parcourir les lignes (en commençant par la ligne 2, après les headers)
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not row[0]:  # Si la première colonne est vide, on arrête
                    break
                
                results["total"] += 1
                
                # Extraire les données
                nom = row[0]
                prenom = row[1]
                date_naissance = row[2]
                genre = row[3]
                contact = row[4]
                projet_name = row[5]
                engagement_name = row[6]
                poste_nom = row[7]
                categorie_poste = row[8]
                type_contrat = row[9]
                type_personne = row[10]
                poste = row[11]
                diplome = row[12]
                ecole = row[13]
                date_debut = row[14]
                date_fin = row[15]
                region = row[16]
                departement = row[17]
                sous_prefecture = row[18]
                
                # Validation basique
                if not nom or not prenom or not projet_name or not poste_nom or not date_debut:
                    results["errors"].append({
                        "row": row_num,
                        "error": "Colonnes obligatoires manquantes (Nom, Prénom, Projet, Poste, Date Début)"
                    })
                    continue
                
                # Récupérer le projet si spécifié
                projet_id_to_use = projet_id
                if projet_name:
                    projet = db.query(Projet).filter(Projet.nom == projet_name).first()
                    if not projet:
                        results["errors"].append({
                            "row": row_num,
                            "error": f"Projet '{projet_name}' non trouvé"
                        })
                        continue
                    projet_id_to_use = projet.id
                
                # Récupérer l'engagement si spécifié
                engagement_id_to_use = None
                if engagement_name:
                    engagement = db.query(Engagement).filter(Engagement.nom == engagement_name).first()
                    if not engagement:
                        results["errors"].append({
                            "row": row_num,
                            "error": f"Engagement '{engagement_name}' non trouvé"
                        })
                        continue
                    engagement_id_to_use = engagement.id
                
                # Créer l'employé
                fic_personne_id = str(uuid.uuid4())
                fic_personne = FicPersonne(
                    id=fic_personne_id,
                    nom=nom,
                    prenom=prenom,
                    date_naissance=date_naissance,
                    genre=genre,
                    contact=contact
                )
                db.add(fic_personne)
                db.flush()  # Flush pour obtenir l'ID
                
                # Créer le contrat
                contrat_id = str(uuid.uuid4())
                contrat = Contrat(
                    id=contrat_id,
                    fic_personne_id=fic_personne_id,
                    projet_id=projet_id_to_use,
                    engagement_id=engagement_id_to_use,
                    poste_nom=poste_nom,
                    categorie_poste=categorie_poste,
                    type_contrat=type_contrat,
                    type_personne=type_personne,
                    poste=poste,
                    date_debut=date_debut,
                    date_fin=date_fin,
                    diplome=diplome,
                    ecole=ecole
                )
                db.add(contrat)
                db.flush()  # Flush pour obtenir l'ID
                
                # Créer la localisation si les données sont fournies
                if region or departement or sous_prefecture:
                    localisation_id = str(uuid.uuid4())
                    localisation = FicPersonneLocalisation(
                        id=localisation_id,
                        contrat_id=contrat_id,
                        region_id=region,
                        departement_id=departement,
                        sous_prefecture_id=sous_prefecture
                    )
                    db.add(localisation)
                
                results["success"] += 1
                results["employees"].append({
                    "nom": nom,
                    "prenom": prenom,
                    "poste": poste_nom
                })
                
            except Exception as e:
                results["errors"].append({
                    "row": row_num,
                    "error": str(e)
                })
        
        # Commit les changements
        db.commit()
        
        return results
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/export-employees")
async def export_employees(acteur_id: str = None, db: Session = Depends(get_db)):
    """Exporter les employés en CSV avec les engagements"""
    
    try:
        import csv
        from io import StringIO
        
        # Créer un StringIO pour le CSV
        output = StringIO()
        
        # Définir les colonnes
        fieldnames = [
            "ID",
            "Nom",
            "Prénom",
            "Date de Naissance",
            "Genre",
            "Contact",
            "Matricule",
            "Projet",
            "Engagement",
            "Poste Nom",
            "Catégorie Poste",
            "Type Contrat",
            "Type Personne",
            "Poste",
            "Diplôme",
            "École",
            "Date Début Contrat",
            "Date Fin Contrat",
            "Région",
            "Département",
            "Sous-Préfecture"
        ]
        
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        
        # Récupérer les employés
        query = db.query(FicPersonne)
        if acteur_id:
            # Récupérer les employés via les contrats
            employees = db.query(FicPersonne).join(Contrat).all()
        else:
            employees = db.query(FicPersonne).all()
        
        # Écrire les données
        for employee in employees:
            # Récupérer le contrat de l'employé
            contrat = db.query(Contrat).filter(Contrat.fic_personne_id == employee.id).first()
            
            if contrat:
                # Récupérer le projet
                projet = db.query(Projet).filter(Projet.id == contrat.projet_id).first() if contrat.projet_id else None
                
                # Récupérer l'engagement
                engagement = db.query(Engagement).filter(Engagement.id == contrat.engagement_id).first() if contrat.engagement_id else None
                
                # Récupérer la localisation
                localisation = db.query(FicPersonneLocalisation).filter(FicPersonneLocalisation.contrat_id == contrat.id).first()
                
                writer.writerow({
                    "ID": employee.id,
                    "Nom": employee.nom,
                    "Prénom": employee.prenom,
                    "Date de Naissance": employee.date_naissance or "",
                    "Genre": employee.genre or "",
                    "Contact": employee.contact or "",
                    "Matricule": employee.matricule or "",
                    "Projet": projet.nom if projet else "",
                    "Engagement": engagement.nom if engagement else "",
                    "Poste Nom": contrat.poste_nom or "",
                    "Catégorie Poste": contrat.categorie_poste or "",
                    "Type Contrat": contrat.type_contrat or "",
                    "Type Personne": contrat.type_personne or "",
                    "Poste": contrat.poste or "",
                    "Diplôme": contrat.diplome or "",
                    "École": contrat.ecole or "",
                    "Date Début Contrat": contrat.date_debut or "",
                    "Date Fin Contrat": contrat.date_fin or "",
                    "Région": localisation.region_id if localisation else "",
                    "Département": localisation.departement_id if localisation else "",
                    "Sous-Préfecture": localisation.sous_prefecture_id if localisation else "",
                })
        
        # Préparer la réponse
        output.seek(0)
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=employes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"}
        )
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
